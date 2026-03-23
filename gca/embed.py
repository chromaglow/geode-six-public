"""
Geode Six — GCA Embedding Module
Handles text extraction, chunking, and Chroma vector DB indexing.

Embedding strategy:
  Primary: nomic-embed-text via Ollama
  Fallback: sentence-transformers/all-MiniLM-L6-v2

v2: Two-tier metadata (tier, code, version, excerpt).
"""

import logging
import os
import re
import uuid as uuid_mod
from pathlib import Path
from typing import Optional

import httpx
from dotenv import load_dotenv

load_dotenv()

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
OLLAMA_HOST = os.getenv("OLLAMA_HOST", "http://localhost:11434")
CHROMA_PATH = os.getenv("CHROMA_PATH", "/mnt/nvme/chroma")
GCA_ROOT = os.getenv("GCA_ROOT", "/mnt/nvme/gca")
LOG_PATH = os.getenv("LOG_PATH", "/mnt/nvme/logs")

EMBED_MODEL = "nomic-embed-text"
CHUNK_SIZE = 500  # characters per chunk
CHUNK_OVERLAP = 50  # overlap between chunks
COLLECTION_NAME = "gca_documents"

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
os.makedirs(LOG_PATH, exist_ok=True)
embed_logger = logging.getLogger("geode.embed")
embed_logger.setLevel(logging.INFO)
_fh = logging.FileHandler(os.path.join(LOG_PATH, "embed.log"))
_fh.setFormatter(logging.Formatter("%(asctime)s %(message)s"))
embed_logger.addHandler(_fh)

# ---------------------------------------------------------------------------
# Chroma client (lazy init)
# ---------------------------------------------------------------------------
_chroma_client = None
_collection = None
_use_ollama_embeddings = True


def _get_collection():
    """Get or create the Chroma collection."""
    global _chroma_client, _collection, _use_ollama_embeddings

    if _collection is not None:
        return _collection

    import chromadb

    os.makedirs(CHROMA_PATH, exist_ok=True)
    _chroma_client = chromadb.PersistentClient(path=CHROMA_PATH)

    # Try Ollama embeddings first
    try:
        _collection = _chroma_client.get_or_create_collection(
            name=COLLECTION_NAME,
            metadata={"hnsw:space": "cosine"},
        )
        _use_ollama_embeddings = True
        embed_logger.info("Using Chroma collection with Ollama embeddings")
    except Exception as e:
        embed_logger.error(f"Chroma init error: {e}")
        raise

    return _collection


def drop_collection():
    """Drop the Chroma collection entirely (for migration)."""
    global _chroma_client, _collection

    import chromadb

    os.makedirs(CHROMA_PATH, exist_ok=True)
    _chroma_client = chromadb.PersistentClient(path=CHROMA_PATH)

    try:
        _chroma_client.delete_collection(name=COLLECTION_NAME)
        embed_logger.info(f"Dropped collection: {COLLECTION_NAME}")
    except Exception:
        embed_logger.info(f"Collection {COLLECTION_NAME} did not exist, nothing to drop")

    _collection = None


# ---------------------------------------------------------------------------
# Embedding via Ollama
# ---------------------------------------------------------------------------


async def get_embedding(text: str) -> list[float]:
    """Get embedding vector from Ollama's nomic-embed-text."""
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.post(
                f"{OLLAMA_HOST}/api/embeddings",
                json={"model": EMBED_MODEL, "prompt": text},
            )
            resp.raise_for_status()
            return resp.json()["embedding"]
    except Exception as e:
        embed_logger.warning(f"Ollama embedding failed, trying fallback: {e}")
        return _fallback_embedding(text)


def _fallback_embedding(text: str) -> list[float]:
    """Fallback: use sentence-transformers if Ollama embedding unavailable."""
    try:
        from sentence_transformers import SentenceTransformer
        model = SentenceTransformer("all-MiniLM-L6-v2")
        embedding = model.encode(text).tolist()
        embed_logger.info("Using sentence-transformers fallback for embedding")
        return embedding
    except ImportError:
        embed_logger.error(
            "sentence-transformers not installed. "
            "Install with: pip install sentence-transformers"
        )
        raise


# ---------------------------------------------------------------------------
# Text extraction (reuse from intake)
# ---------------------------------------------------------------------------


def extract_full_text(filepath: str) -> str:
    """Extract full text from a supported file."""
    ext = os.path.splitext(filepath)[1].lower()

    try:
        if ext == ".pdf":
            from PyPDF2 import PdfReader
            reader = PdfReader(filepath)
            return "\n".join(page.extract_text() or "" for page in reader.pages)

        elif ext == ".docx":
            from docx import Document
            doc = Document(filepath)
            return "\n".join(p.text for p in doc.paragraphs)

        elif ext in (".txt", ".md"):
            with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()

        elif ext == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(" | ".join(str(c) for c in row if c))
            return "\n".join(rows)

        elif ext in (".jpg", ".jpeg", ".png"):
            return f"[Image: {os.path.basename(filepath)}]"

    except Exception as e:
        embed_logger.error(f"Text extraction failed for {filepath}: {e}")
    return ""


# ---------------------------------------------------------------------------
# Chunking
# ---------------------------------------------------------------------------


def chunk_text(text: str, chunk_size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> list[str]:
    """Split text into overlapping chunks."""
    if not text or len(text) <= chunk_size:
        return [text] if text else []

    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end]
        if chunk.strip():
            chunks.append(chunk)
        start = end - overlap

    return chunks


# ---------------------------------------------------------------------------
# Filename metadata parser
# ---------------------------------------------------------------------------


def _parse_filename_metadata(filename: str) -> dict:
    """Parse GCA metadata from a filename."""
    parts = filename.split("_")
    metadata = {
        "filename": filename,
        "code": parts[0] if len(parts) > 0 else "",
        "type": parts[1] if len(parts) > 1 else "",
        "description": parts[2] if len(parts) > 2 else "",
    }
    # Extract date
    date_match = re.search(r"_(\d{8})_", filename)
    if date_match:
        metadata["date"] = date_match.group(1)
    # Extract version
    version_match = re.search(r"_v(\d+\.\d+)", filename)
    if version_match:
        metadata["version"] = version_match.group(1)
    return metadata


# ---------------------------------------------------------------------------
# Embedding functions
# ---------------------------------------------------------------------------


async def embed_file(
    filepath: str,
    tier: str = "",
    project: str = "",
    type_code: str = "",
    description: str = "",
    date: str = "",
    version: str = "",
):
    """Embed a single file into Chroma with v2 metadata schema."""
    collection = _get_collection()
    filename = os.path.basename(filepath)

    # Extract text
    text = extract_full_text(filepath)
    if not text:
        embed_logger.warning(f"No text extracted from {filename}, skipping")
        return

    # Chunk
    chunks = chunk_text(text)
    if not chunks:
        embed_logger.warning(f"No chunks produced from {filename}, skipping")
        return

    # If metadata not provided, try to parse from filename
    if not project:
        parsed = _parse_filename_metadata(filename)
        project = parsed.get("code", "")
        type_code = type_code or parsed.get("type", "")
        description = description or parsed.get("description", "")
        date = date or parsed.get("date", "")
        version = version or parsed.get("version", "")

    # If tier not provided, try to resolve from codes
    if not tier:
        try:
            from gca.codes import tier_for_code
            tier = tier_for_code(project) or ""
        except Exception:
            tier = ""

    # Build excerpt (first 200 chars of full text)
    excerpt = text[:200].strip()

    # Document ID
    doc_id = str(uuid_mod.uuid4())

    # Remove existing entries for this file (re-embed)
    try:
        existing = collection.get(where={"file_path": filepath})
        if existing and existing["ids"]:
            collection.delete(ids=existing["ids"])
    except Exception:
        pass

    # Embed each chunk
    ids = []
    embeddings = []
    documents = []
    metadatas = []

    for i, chunk in enumerate(chunks):
        chunk_id = f"{filename}_chunk_{i}"
        embedding = await get_embedding(chunk)

        ids.append(chunk_id)
        embeddings.append(embedding)
        documents.append(chunk)
        metadatas.append({
            "doc_id": doc_id,
            "filename": filename,
            "tier": tier,
            "code": project,
            "type": type_code,
            "date": date,
            "version": version,
            "file_path": filepath,
            "excerpt": excerpt,
            "chunk_index": str(i),
        })

    # Add to Chroma
    collection.add(
        ids=ids,
        embeddings=embeddings,
        documents=documents,
        metadatas=metadatas,
    )

    embed_logger.info(f"Embedded {filename}: {len(chunks)} chunks")


async def embed_all_files():
    """Embed all files in GCA_ROOT into Chroma (full re-index).
    Iterates the two-tier structure: GCA_ROOT/[tier]/[code]/[files].
    """
    embed_logger.info(f"Starting full embedding pass over {GCA_ROOT}")
    count = 0

    for tier in ("Projects", "Operations"):
        tier_path = os.path.join(GCA_ROOT, tier)
        if not os.path.isdir(tier_path):
            continue

        for code_dir in os.listdir(tier_path):
            code_path = os.path.join(tier_path, code_dir)
            if not os.path.isdir(code_path):
                continue

            for filename in os.listdir(code_path):
                filepath = os.path.join(code_path, filename)
                if not os.path.isfile(filepath):
                    continue

                parsed = _parse_filename_metadata(filename)
                try:
                    await embed_file(
                        filepath,
                        tier=tier,
                        project=parsed.get("code", code_dir),
                        type_code=parsed.get("type", ""),
                        description=parsed.get("description", ""),
                        date=parsed.get("date", ""),
                        version=parsed.get("version", ""),
                    )
                    count += 1
                except Exception as e:
                    embed_logger.error(f"Failed to embed {filename}: {e}")

    embed_logger.info(f"Full embedding pass complete: {count} files embedded")
    return count
